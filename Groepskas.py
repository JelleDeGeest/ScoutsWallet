import tkinter
import customtkinter
import csv
import openpyxl
import copy
import os
import pandas as pd
import win32com.client
import TopLevelWindows as tlw
import time
from CTkMessagebox import CTkMessagebox

class GroepsKas:
    def __init__(self, root, mainmenu):
        self.root = root
        self.mainmenu = mainmenu
        self.current_path = os.getcwd()
        self.tabladen = {}
        self.loadfiles()
        self.frame = customtkinter.CTkFrame(master=self.root)
        self.groeps_frame = self.groepskasframe()
        self.groeps_frame.pack(pady=0, padx=0, fill="both", expand=True)

    def loadfiles(self):
        self.lastchecked = self.mainmenu.config["laatste_transactie"]
        self.huidig_jaar = self.mainmenu.config["huidig_jaar"]
        self.years = [f.split()[1][:4] for f in os.listdir(os.path.join(self.current_path, "Files") ) if f.endswith('.xlsx') and f.startswith('Groepskas')]
        for jaar in self.years:
            wb = openpyxl.load_workbook(os.path.join(self.current_path, "Files", "Groepskas " + jaar + ".xlsx"))
            self.tabladen[jaar] = [f for f in wb.sheetnames if f != "Algemeen" and f != "Sjabloon"]
        self.years.sort(reverse=True)
        
        print(self.current_path)
        print(self.tabladen)
        print(self.years)

    
    def groepskasframe(self):
        frame = customtkinter.CTkFrame(master=self.frame)
        label = customtkinter.CTkLabel(master=frame, text="Groepskas", font=("Arial", 30))
        label.grid(row = 0, column = 0, columnspan = 2, pady=12, padx=10)
        
        addtransactionsbutton = customtkinter.CTkButton(master=frame, text="Voeg transacties toe", command=self.clickaddtransactions)
        addtransactionsbutton.grid(row = 1, column = 1, columnspan = 1, sticky = "", pady=(0,24), padx=10,)

        newyearbutton = customtkinter.CTkButton(master=frame, text="Nieuw jaar", command=self.clicknewyear)
        newyearbutton.grid(row = 1, column = 0, sticky = "", pady=(0,24), padx=10,)

        frame.columnconfigure(1, weight=5)
        frame.rowconfigure(2, weight=1)
        frame.columnconfigure(0, weight=1)

        self.overzicht_frame = Overzicht(frame, self)
        
        return frame
    
    def clicknewyear(self):

        last_year = self.years[0]
        new_year = str(last_year[len(last_year)//2:]) + str(int(last_year[len(last_year)//2:]) + 1)

        #Get the previous year end balance to put it in the start of the new one
        wb = openpyxl.load_workbook(os.path.join(self.current_path, "Files", "Groepskas " + last_year + ".xlsx"), data_only=True)
        previous_year_end_balance = wb["Algemeen"]['D14'].value

        #Open the template workbook and save it under the new name
        wb = openpyxl.load_workbook(os.path.join(self.current_path, "template.xlsx"))
        wb["Algemeen"]['B2'].value = previous_year_end_balance


        wb.save(os.path.join(self.current_path, "Files", "Groepskas " + new_year + ".xlsx"))

        self.loadfiles()
        self.overzicht_frame.loadyears()
        self.mainmenu.evaluateworkbooks()


    def getgroepskasframe(self):
        return self.frame

    def clickaddtransactions(self):
        self.addtransactions_frame = AddTransactions(self)
        self.groeps_frame.pack_forget()
        self.addtransactions_frame.getframe().pack(pady=0, padx=0, fill="both", expand=True)

    def returntogroepskasframe(self):
        self.addtransactions_frame.getframe().pack_forget()
        self.groeps_frame.pack(pady=0, padx=0, fill="both", expand=True)
        self.loadfiles()

class Overzicht:
    def __init__(self, frame, groepskas):
        # Create an oversight frame which is split up in selectframe (left) and dataframe (right)
        self.frame = frame
        self.groepskas = groepskas
        self.selectframe = customtkinter.CTkScrollableFrame(master=self.frame, bg_color="red")
        self.selectframe.grid(row = 2, column = 0, sticky = "nswe", pady=0, padx=0)
        self.dataframe = customtkinter .CTkScrollableFrame(master=self.frame, bg_color = "blue")
        self.dataframe.grid(row = 2, column = 1, sticky = "nswe", pady=0, padx=0)

        # v1 = customtkinter.CTkScrollbar(self.selectframe, command=self.selectframe.yview)
        # v2 = customtkinter.CTkScrollbar(self.dataframeframe, command=self.dataframeframe.yview)

        #Create and load in the different years
        self.loadyears()
    
    def loadyears(self):
        
        # Create a list of buttons for the years
        self.yearbuttons = []
        self.selectframe.columnconfigure(0, weight=1)
        for i in range(len(self.groepskas.years)):
            self.yearbuttons.append(customtkinter.CTkButton(master=self.selectframe, text=self.groepskas.years[i], command= lambda i=i: self.clickyear(self.groepskas.years[i])))
            self.yearbuttons[i].grid(row = i, column = 0, sticky = "nswe", pady=4, padx=2)
            # self.yearbuttons[i].bind("<Button-1>", self.clickyear)

    def clickyear(self, year):
        # Fetch data from the excel file
        wb = openpyxl.load_workbook(os.path.join(self.groepskas.current_path, "Files", "Groepskas " + year + ".xlsx"), data_only=True)
        sheet = wb["Algemeen"]
        
        # Read data from the excel file
        data = []
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                data.append((row[0], round(row[1], 2))) 

        # Clear the dataframe
        for widget in self.dataframe.winfo_children():
            widget.destroy()

        # Create 2 labels for each transaction, 1 with the name and 1 with the amount, if the amount is negative it needs to be red otherwise green
        for i in range(len(data)):
            name = customtkinter.CTkLabel(master=self.dataframe, text=data[i][0])
            name.grid(row=i, column=0, sticky="nsw", pady=4, padx=10)
            amount = customtkinter.CTkLabel(master=self.dataframe, text=data[i][1])
            amount.grid(row=i, column=1, sticky="nsw", pady=4, padx=10)
            if data[i][1] < 0:
                amount.configure(text_color="red")
            else:
                amount.configure(text_color="green")
        # self.dataframe.columnconfigure(0, weight=1)
        
        
            


class AddTransactions:
    def __init__(self, groepskas):
        self.groepskas = groepskas
        self.add_frame = customtkinter.CTkFrame(master=self.groepskas.frame, bg_color="red")
        label = customtkinter.CTkLabel(master=self.add_frame, text="Transacties toevoegen", font=("Arial", 30))
        label.grid(row = 0, column = 0, sticky="nswe", pady=8, padx=10)
        self.add_frame.columnconfigure(0, weight=1)
        self.add_frame.rowconfigure(2, weight=1)

        self.choosefileframe = customtkinter.CTkFrame(master=self.add_frame, fg_color="transparent")
        self.choosefileframe.grid(row = 1, column = 0, columnspan = 1, sticky = "", pady=(0,0), padx=10)
        self.choosefilebutton = customtkinter.CTkButton(master=self.choosefileframe, text="Bestand kiezen", command=self.choosefileclicked)
        self.choosefilebutton.grid(row = 0, column = 0, columnspan = 1, sticky = "", pady=(0,12), padx=10,)
        self.backbutton2 = customtkinter.CTkButton(master=self.choosefileframe, text="Terug", command=self.groepskas.returntogroepskasframe)
        self.backbutton2.grid(row = 0, column = 1, columnspan = 1, sticky = "", pady=(0,12), padx=10)
        self.buttonframe = customtkinter.CTkFrame(master=self.add_frame, fg_color="transparent")
        self.savebutton = customtkinter.CTkButton(master=self.buttonframe, text="Opslaan", command=self.save_data_preparation) 
        self.savebutton.grid(row = 1, column = 0, columnspan = 2, sticky = "", pady=(0,12), padx=10)
        self.newtabbutton = customtkinter.CTkButton(master=self.buttonframe, text="Nieuw tablad", command=self.newtab)
        self.newtabbutton.grid(row = 0, column = 1, columnspan = 1, sticky = "", pady=(0,12), padx=10)
        self.backbutton = customtkinter.CTkButton(master=self.buttonframe, text="Terug", command=self.groepskas.returntogroepskasframe)
        self.backbutton.grid(row = 0, column = 0, columnspan = 1, sticky = "", pady=(0,12), padx=10)
    
    def save_data_preparation(self):   
         waitwindow = tlw.WaitWindow(self, master=self.groepskas.mainmenu.root)
        #  waitwindow.after(100, waitwindow.lift)
    
    def savedata(self):
        
        #Check if all the data is filled in
        if not self.checksave():
            return
        #Check if the files are writable
        if not self.checkwritability():
            CTkMessagebox(title="Error", message="Geen permissie om bestanden aan te passen. Check of 1 van de excel bestanden of het config bestand nog ergens geopend zijn.", icon="cancel")
            return
        

        #Save the data to the excel files
        arrays = [row for row in zip([row[0] for row in self.transactions[1:]], [row[1] for row in self.transactions[1:]], [row[2] for row in self.transactions[1:]], [t.cget("text") for t in self.names], [t.cget("text") for t in self.descriptions], [t.get() for t in self.jaar_option_vars], [t.get() for t in self.tablad_option_vars])]
        df = pd.DataFrame(arrays, columns=["Datum", "Bedrag", "Saldo", "Naam", "Beschrijving", "Jaar", "Tablad"])
        grouped_byyear = df.groupby("Jaar")
        changed_years = []
        for year, group in grouped_byyear:
            changed_years.append(int(year))
            grouped_bytablad = group.groupby("Tablad")
            file_path = self.groepskas.current_path + "/Files/Groepskas " + year + ".xlsx"
            wb = openpyxl.load_workbook(file_path)
            for tablad, tablad_group in grouped_bytablad:
                sheet = wb[tablad]
                first_empty_row = 1
                while sheet.cell(first_empty_row, 1).value is not None or sheet.cell(first_empty_row, 2).value is not None:
                    first_empty_row += 1
                lenght = len(tablad_group)
                for i in range(lenght):
                    sheet.cell(row=first_empty_row + i, column=1).value = tablad_group['Naam'].iloc[lenght-i-1] + ": " + tablad_group['Beschrijving'].iloc[lenght-i-1]
                    sheet.cell(row=first_empty_row + i, column=2).value = float(tablad_group['Bedrag'].iloc[lenght-i-1])
                    sheet.cell(row=first_empty_row, column=2).number_format = '_-€ * #,##0.00_-;-€ * #,##0.00_-;_-€ * "-"??_-;_-@_-'
            wb.save(file_path)

        #Make the starting budget up to date
        self.groepskas.mainmenu.evaluateworkbooks()
        temp_years = [int(year) for year in self.groepskas.years]
        temp_years.sort()
        print("-------------------")
        print(temp_years)
        print(changed_years)
        for i,year in enumerate(temp_years):
            excel = win32com.client.Dispatch("Excel.Application")
            if year >= min(changed_years) and i > 0:
                file_path = self.groepskas.current_path + "/Files/Groepskas " + str(temp_years[i-1]) + ".xlsx"
                wb = openpyxl.load_workbook(file_path, data_only=True)
                previous_total = wb["Algemeen"]["D14"].value
                file_path = self.groepskas.current_path + "/Files/Groepskas " + str(year) + ".xlsx"
                wb = openpyxl.load_workbook(file_path)
                wb["Algemeen"]["B2"].value = previous_total
                wb.save(file_path)
                wb = excel.Workbooks.Open(file_path)
                wb.Save()
                wb.Close(True)  
            excel.Quit()
                
        


        #Save all transaction in transaction file to keep track of all transaction
        file_path = self.groepskas.current_path + "/Files/Transacties.xlsx"
        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Transactions"]
        lenght = len(df)

        first_empty_row = 1
        while sheet.cell(first_empty_row, 1).value is not None:
            first_empty_row += 1

        for i in range(lenght):
            sheet.cell(row=i+first_empty_row, column=1).value = df['Datum'].iloc[lenght-i-1]
            sheet.cell(row=i+first_empty_row, column=2).value = float(df['Bedrag'].iloc[lenght-i-1])
            sheet.cell(row=i+first_empty_row, column=3).value = df['Saldo'].iloc[lenght-i-1]
            sheet.cell(row=i+first_empty_row, column=4).value = df['Naam'].iloc[lenght-i-1]
            sheet.cell(row=i+first_empty_row, column=5).value = df['Beschrijving'].iloc[lenght-i-1]
            sheet.cell(row=i+first_empty_row, column=6).value = int(df['Jaar'].iloc[lenght-i-1])
            sheet.cell(row=i+first_empty_row, column=7).value = df['Tablad'].iloc[lenght-i-1]

        wb.save(file_path)

        #Change the lastchecked config
        self.groepskas.mainmenu.config["laatste_transactie"] = str(self.transactions[1])

        #Also change it in the config file
        with open("Config.txt", "w") as file:
            for key, value in self.groepskas.mainmenu.config.items():
                file.write(key + "! " + value + "\n")

        self.groepskas.returntogroepskasframe()

    def checkwritability(self):
        writable = True
        for filename in os.listdir(self.groepskas.current_path + "/Files"):
            if filename.endswith('.xlsx') and not filename.startswith('~'):
                file_path = os.path.join(self.groepskas.current_path, "Files" , filename)
                try:
                    # Try to open the file in append mode.
                    with open(file_path, 'a'):
                        print(f"File is writable: {file_path}")
                except PermissionError:
                    print(f"Permission error: Can't write to {file_path}")
                    writable = False
                except FileNotFoundError:
                    print(f"File not found: {file_path}")
                    writable = False

        #Check config file availability
        file_path = os.path.join(self.groepskas.current_path, "Config.txt")
        try:
            with open(file_path, 'a'):
                print(f"File is writable: Config.txt")
        except PermissionError:
            print(f"Permission error: Can't write to Config.txt")
            writable = False
        except FileNotFoundError:
            print(f"File not found: Config.txt")
            writable = False



        return writable

    def getframe(self):
        return self.add_frame
    
    # Activates necessary functions when the choosefilebutton is clicked
    def choosefileclicked(self):
        self.transactions = self.gettransactions()
        # self.choosefilebutton.grid_forget()
        if len(self.transactions) <= 1:
            CTkMessagebox(title="Info", message="Er zijn geen nieuwe betalingen om te verwerken", option_1="OK")
            return
        
        self.buttonframe.grid(row = 1, column = 0, columnspan = 2, sticky = "", pady=(8,8), padx=10)
        self.displaytransactions()
        
    def changeyear(self, index):
        self.tablad_comboboxes[index-1].configure(require_redraw=True, values=self.groepskas.tabladen[self.jaar_option_vars[index-1].get()])
        self.tablad_option_vars[index-1].set("Selecteer een tablad")

    def changetablad(self, index):
        print(index, self.tablad_option_vars[index-1].get())

    def checksave(self):
        ready = True
        for i in range(len(self.transactions)-1):
            self.entry_returned(0, i)
            self.entry_returned(1, i)
            if self.tablad_option_vars[i].get() == "Selecteer een tablad" or self.jaar_option_vars[i].get() == "Selecteer een jaar" or isinstance(self.names[i],customtkinter.CTkEntry) or isinstance(self.descriptions[i],customtkinter.CTkEntry):
                ready = False
        return ready
        

            
    
    # Displays all the transactions in the add_frame
    def displaytransactions(self):
        self.transactions_frame = customtkinter.CTkScrollableFrame(master=self.add_frame, bg_color="Blue")
        self.transactions_frame.grid(row = 2, column = 0, sticky = "nswe", pady=0, padx=0)
        self.transactions_frame.columnconfigure(0, weight=1)
        self.transactions_frame.columnconfigure(1, weight=1)
        self.transactions_frame.columnconfigure(2, weight=1)
        self.transactions_frame.columnconfigure(3, weight=1)
        self.transactions_frame.columnconfigure(4, weight=1)
        self.transactions_frame.columnconfigure(5, weight=1)

        self.tablad_option_vars = []
        self.jaar_option_vars = []
        self.tablad_comboboxes = []
        self.names = []
        self.descriptions = []

        # Create a header
        date = customtkinter.CTkLabel(master=self.transactions_frame, text="Datum")
        date.grid(row = 0, column = 0, sticky = "nswe", pady=4, padx=10)
        amount = customtkinter.CTkLabel(master=self.transactions_frame, text="Bedrag")
        amount.grid(row = 0, column = 1, sticky = "nswe", pady=4, padx=10)
        name = customtkinter.CTkLabel(master=self.transactions_frame, text="Naam")
        name.grid(row = 0, column = 2, sticky = "nswe", pady=4, padx=10)
        description = customtkinter.CTkLabel(master=self.transactions_frame, text="Beschrijving")
        description.grid(row = 0, column = 3, sticky = "nswe", pady=4, padx=10)
        jaar = customtkinter.CTkLabel(master=self.transactions_frame, text="Jaar")
        jaar.grid(row = 0, column = 4, sticky = "nswe", pady=4, padx=10)
        tablad = customtkinter.CTkLabel(master=self.transactions_frame, text="Tablad")
        tablad.grid(row = 0, column = 5, sticky = "nswe", pady=4, padx=10)
        

        print(self.transactions)
        for i in range(1,len(self.transactions)):

            # Create a date label
            date = customtkinter.CTkLabel(master=self.transactions_frame, text=self.transactions[i][0])
            date.grid(row = i, column = 0, sticky = "nswe", pady=4, padx=10)
            

            # Create an amount label
            if float(self.transactions[i][1]) < 0:
                amount = customtkinter.CTkLabel(master=self.transactions_frame, text=self.transactions[i][1], text_color="red")
            else:
                amount = customtkinter.CTkLabel(master=self.transactions_frame, text=self.transactions[i][1], text_color="green")
            amount.grid(row = i, column = 1, sticky = "nswe", pady=4, padx=10)

            # Create a name editable label
            if self.transactions[i][3] != "":
                name = customtkinter.CTkLabel(master=self.transactions_frame, text=self.transactions[i][3])
                name.grid(row = i, column = 2, sticky = "nswe", pady=4, padx=10)
                name.bind("<Button-1>", lambda event, index=i: self.label_clicked(0, index))
            else:
                name = customtkinter.CTkEntry(master=self.transactions_frame)
                name.grid(row = i, column = 2, sticky = "nswe", pady=4, padx=10)
                name.bind("<Return>", lambda event, i=i: self.entry_returned(0, i))

            self.names.append(name)

            # Create a description editable label
            if self.transactions[i][4] != "":
                description = customtkinter.CTkLabel(master=self.transactions_frame, text=self.transactions[i][4])
                description.grid(row = i, column = 3, sticky = "nswe", pady=4, padx=10)
                description.bind("<Button-1>", lambda event, index=i: self.label_clicked(1, index))
            else:
                description = customtkinter.CTkEntry(master=self.transactions_frame)
                description.grid(row = i, column = 3, sticky = "nswe", pady=4, padx=10)
                description.bind("<Return>", lambda event, i=i: self.entry_returned(1, i))

            self.descriptions.append(description)

            # Create a combobox with all the years
            option_var = tkinter.StringVar()
            option_var.set(self.groepskas.huidig_jaar)
            self.jaar_option_vars.append(option_var)       
            jaar_menu = customtkinter.CTkOptionMenu(master=self.transactions_frame, variable=option_var, values=self.groepskas.years, command=lambda x, index=i: self.changeyear(index))
            jaar_menu.grid(row = i, column = 4, sticky = "nswe", pady=4, padx=10)

            # Create a combobox with all the sheetnames
            option_var = tkinter.StringVar()
            option_var.set("Selecteer een tablad")
            tablad_menu = customtkinter.CTkOptionMenu(master=self.transactions_frame, variable=option_var, values=self.groepskas.tabladen[self.groepskas.huidig_jaar], command=lambda x, index=i: self.changetablad(index))
            tablad_menu.grid(row = i, column = 5, sticky = "nswe", pady=4, padx=10)
            self.tablad_option_vars.append(option_var)
            self.tablad_comboboxes.append(tablad_menu)

    def label_clicked(self, type, i):
        if type == 0:
            change_list = self.names
        elif type == 1:
            change_list = self.descriptions

        change_list[i-1].grid_forget()
        entry = customtkinter.CTkEntry(master=self.transactions_frame)
        entry.insert(0, change_list[i-1].cget("text"))
        change_list[i-1] = entry
        change_list[i-1].grid(row = i, column = type + 2, sticky = "nswe", pady=4, padx=10)
        entry.bind("<Return>", lambda event, type=type, i=i: self.entry_returned(type, i))

    def entry_returned(self, type, i):
        if type == 0:
            change_list = self.names
        elif type == 1:
            change_list = self.descriptions
        
        if isinstance(change_list[i], customtkinter.CTkEntry) and change_list[i].get() != "":
            change_list[i].grid_forget()
            label = customtkinter.CTkLabel(master=self.transactions_frame, text=change_list[i].get())
            label.grid(row = i+1, column = type + 2, sticky = "nswe", pady=4, padx=10)
            label.bind("<Button-1>", lambda event, type=type, i=i: self.label_clicked(type, i+1))
            change_list[i] = label
        

    # Gets all the transactions to add and stores them in a list
    def gettransactions(self):
        filelist = customtkinter.filedialog.askopenfilename(title ='Select Transactions', filetypes=[('CSV files', '*.csv')])
        if(filelist == ""):
            print("Invalid Transaction Document")
            exit()
        transactions = []
        with open(filelist, 'r') as file:
            csvreader = csv.reader(file)
            temp1 = self.groepskas.lastchecked.split(",")
            print(temp1)
            for row in csvreader:
                temp_transaction = ".".join(row).split(';')
                transaction = [temp_transaction[5],temp_transaction[8],temp_transaction[9],temp_transaction[14].replace('                                                                       ', ""), temp_transaction[17].replace("                                                                                                                                            ", ""), "",""] 
                temp2 = str(transaction).split(",")
                if (temp1[0] == temp2[0] and temp1[3] == temp2[3] and temp1[4] == temp2[4]) or (transaction[0] == "" and transaction[1] == ""):
                    break
                else:
                    transactions.append(transaction)
            return transactions
    
    def newtab(self):
        newtabwindow = tlw.NewTabWindow(self, master=self.groepskas.mainmenu.root)
        newtabwindow.after(100, newtabwindow.lift)


    
    