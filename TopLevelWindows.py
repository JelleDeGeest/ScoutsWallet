import customtkinter
import tkinter
import openpyxl
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from CTkMessagebox import CTkMessagebox



class NewTabWindow(customtkinter.CTkToplevel):
    def __init__(self, add_transactions, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.add_transactions = add_transactions
        self.geometry("500x350")

        self.setup_window()

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.temp=self.master.winfo_geometry()
        self.geometry(self.temp)
        self.master.withdraw()


    def destroy(self):
        super().destroy()
        self.master.deiconify()
        self.master.geometry(self.temp)
    
    def setup_window(self):
        label_year = customtkinter.CTkLabel(master=self, text="Kies in welk jaar je een tablad wilt toevoegen")
        year_option_var = tkinter.StringVar()
        year_option_var.set(self.add_transactions.groepskas.huidig_jaar)
        year_menu = customtkinter.CTkOptionMenu(master=self, variable=year_option_var, values=self.add_transactions.groepskas.years)
        label_year.grid(row=0, column=0, sticky="", padx=24, pady=(16,0))
        year_menu.grid(row=1, column=0, sticky="", padx=24, pady=(4,0))
        self.columnconfigure(0, weight=1)

        label_tablad = customtkinter.CTkLabel(master=self, text="Kies een naam voor het tablad:")
        tablad_entry = customtkinter.CTkEntry(master=self, width=300)
        label_tablad.grid(row=2, column=0, sticky="", padx=24, pady=(16,0))
        tablad_entry.grid(row=3, column=0, sticky="", padx=24, pady=(4,0))

        label_afkorting = customtkinter.CTkLabel(master=self, text="Kies een 3 letter afkorting voor het tablad voor overschrijvingen:")
        label_afkorting2 = customtkinter.CTkLabel(master=self, text="Bvb: LFW voor leefweek")
        afkorting_entry = customtkinter.CTkEntry(master=self, width=300)
        label_afkorting.grid(row=4, column=0, sticky="", padx=24, pady=(16,0))
        label_afkorting2.grid(row=5, column=0, sticky="", padx=24, pady=(0,0))
        afkorting_entry.grid(row=6, column=0, sticky="", padx=24, pady=(4,0))

        label_payconiq = customtkinter.CTkLabel(master=self, text="Naam van de payconiq webshop/inschrijvingen voor automatische verwerking:")
        label_payconiq2 = customtkinter.CTkLabel(master=self, text="(Laat leeg als er geen is)")
        payconiq_entry = customtkinter.CTkEntry(master=self, width=300)
        label_payconiq.grid(row=7, column=0, sticky="", padx=24, pady=(16,0))
        label_payconiq2.grid(row=8, column=0, sticky="", padx=24, pady=(0,0))
        payconiq_entry.grid(row=9, column=0, sticky="", padx=24, pady=(4,0))
        
        button = customtkinter.CTkButton(master=self, text="Toevoegen", command=lambda: self.add_tablad(year_option_var.get(), tablad_entry.get(), afkorting_entry.get(), payconiq_entry.get()))
        button.grid(row=10, column=0, sticky="", padx=24, pady=(16,0))

    def add_tablad(self, year, tablad, afkorting, payconiq):
        
        #Check if the filled in values are valid
        if tablad in self.add_transactions.groepskas.tabladen[year]:
            CTkMessagebox(title="Error", message="Tablad bestaat al!", icon="cancel")
            return
        
        if len(afkorting) != 3:
            CTkMessagebox(title="Error", message="Afkorting moet 3 letters lang zijn!", icon="cancel")
            return
        
        afkortingen = []
        file_path = os.path.join(self.add_transactions.groepskas.current_path, "Files", "Groepskas " + year + ".xlsx")
        wb = openpyxl.load_workbook(file_path)
        if tablad in wb.sheetnames:
            CTkMessagebox(title="Error", message="Tablad bestaat al!", icon="cancel")
            return
        for sheet_name in wb.sheetnames:
            if sheet_name != "Algemeen" and sheet_name != "Sjabloon":
                afkortingen.append(wb[sheet_name]["E4"].value)
        if afkorting in afkortingen:
            CTkMessagebox(title="Error", message="Afkorting bestaat al!", icon="cancel")
            return
        
        #Check if file is available for editing
        if not os.access(file_path, os.W_OK):
            CTkMessagebox(title="Error", message="Excel bestand is niet bewerkbaar zorg ervoor dat je het niet meer open hebt staan!", icon="cancel")
            return

        #Edit files
        source_sheet = wb["Sjabloon"]
        copied_sheet = wb.copy_worksheet(source_sheet)
        copied_sheet.title = tablad
        copied_sheet["E4"] = afkorting
        copied_sheet["E5"] = payconiq

        # Create a CellIsRule for green fill and font
        green_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
        green_font = Font(color="006100")  # Black font
        green_rule = CellIsRule(operator="greaterThan", formula=[0], fill=green_fill, font=green_font)

        # Create a CellIsRule for red fill and font
        red_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
        red_font = Font(color="9C0006")  # Black font
        red_rule = CellIsRule(operator="lessThan", formula=[0], fill=red_fill, font=red_font)


        # Apply the rules to a range of cells
        copied_sheet.conditional_formatting.add("B2:B1000", green_rule)
        copied_sheet.conditional_formatting.add("B2:B1000", red_rule)
    
        #Add the added tab in the Algemeen tablad
        sheet = wb["Algemeen"]
        row = 1
        while sheet["A" + str(row)].value != None:
            row += 1
        sheet["A" + str(row)].hyperlink = "#" + tablad + "!A1"
        sheet["A" + str(row)].value = tablad
        sheet["B" + str(row)].value = "='" + tablad +"'!E3"

        wb.save(file_path)

        #Change the variables in groepskas
        self.add_transactions.groepskas.loadfiles()
        for i in range(len(self.add_transactions.tablad_comboboxes)):
            if self.add_transactions.jaar_option_vars[i].get() == year:
                self.add_transactions.tablad_comboboxes[i].configure(values=self.add_transactions.groepskas.tabladen[year])

        self.destroy()
    
class WaitWindow(customtkinter.CTkToplevel):
    def __init__(self, add_transactions, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.center_window()
        self.setup_window()

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.temp = self.master.winfo_geometry()
        self.master.withdraw()

    def destroy(self):
        return

    def center_window(self):
        # Get screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calculate position x and y coordinates
        x = (screen_width // 2) - (500 // 2)
        y = (screen_height // 2) - (350 // 2)
        
        self.geometry(f"{500}x{350}+{x}+{y}")

    
    def setup_window(self):
        label = customtkinter.CTkLabel(master=self, text="Bestanden worden opgeslagen, even geduld dit kan even duren")
        label.grid(row=0, column=0, sticky="", padx=24, pady=(0,0))
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

    def finish(self):
        super().destroy()
        self.master.deiconify()
        self.master.geometry(self.temp)